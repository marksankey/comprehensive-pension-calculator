"""
Enhanced Excel Export Module for Bond Strategy Calculator
=========================================================

This module extends the existing Excel export functionality with additional
practical sheets for daily bond investment management.

Usage:
    from excel_enhanced_export import EnhancedExcelExporter
    
    exporter = EnhancedExcelExporter()
    excel_data = exporter.create_complete_excel_export(
        annual_data, sipp_ladder, isa_ladder, scenario_results
    )
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from io import BytesIO
import logging
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows

class EnhancedExcelExporter:
    """Enhanced Excel export functionality for bond investment tools"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # Define color schemes
        self.colors = {
            'header': 'FFD700',      # Gold
            'sipp': 'E6F3FF',        # Light Blue
            'isa': 'E6FFE6',         # Light Green
            'warning': 'FFE6E6',     # Light Red
            'good': 'E6FFE6',        # Light Green
            'neutral': 'FFF5E6',     # Light Orange
            'border': '000000'       # Black
        }
        
        # Excel formatting styles
        self.styles = {
            'header': Font(bold=True, size=12),
            'subheader': Font(bold=True, size=10),
            'currency': '#,##0.00',
            'percentage': '0.00%',
            'date': 'DD/MM/YYYY'
        }
    
    def create_complete_excel_export(self, annual_data: List[Dict], 
                                   sipp_ladder: pd.DataFrame, 
                                   isa_ladder: pd.DataFrame,
                                   scenario_results: Optional[Dict] = None,
                                   monte_carlo_results: Optional[List] = None) -> bytes:
        """
        Create comprehensive Excel export with all sheets
        
        Args:
            annual_data: List of annual simulation results
            sipp_ladder: SIPP bond ladder DataFrame
            isa_ladder: ISA bond ladder DataFrame
            scenario_results: Optional scenario analysis results
            monte_carlo_results: Optional Monte Carlo simulation results
            
        Returns:
            Excel file as bytes
        """
        try:
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Create all sheets
                self._create_annual_analysis_sheet(writer, annual_data)
                self._create_bond_price_tracker_sheet(writer, sipp_ladder, isa_ladder)
                self._create_purchase_log_sheet(writer)
                self._create_quick_calculator_sheet(writer)
                self._create_portfolio_summary_sheet(writer, annual_data, sipp_ladder, isa_ladder)
                self._create_risk_dashboard_sheet(writer, sipp_ladder, isa_ladder)
                self._create_yield_analysis_sheet(writer, sipp_ladder, isa_ladder)
                self._create_implementation_checklist_sheet(writer)
                
                # Include original sheets if data provided
                if sipp_ladder is not None and not sipp_ladder.empty:
                    self._create_sipp_ladder_sheet(writer, sipp_ladder)
                if isa_ladder is not None and not isa_ladder.empty:
                    self._create_isa_ladder_sheet(writer, isa_ladder)
                
                # Optional advanced sheets
                if scenario_results:
                    self._create_scenario_analysis_sheet(writer, scenario_results)
                if monte_carlo_results:
                    self._create_monte_carlo_sheet(writer, monte_carlo_results)
                
                # Apply formatting to all sheets
                self._apply_formatting(writer)
            
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error creating enhanced Excel export: {str(e)}")
            raise
    
    def _create_bond_price_tracker_sheet(self, writer: pd.ExcelWriter, 
                                       sipp_ladder: pd.DataFrame, 
                                       isa_ladder: pd.DataFrame):
        """Create daily bond price tracking sheet"""
        
        # Combine bond data from both ladders
        tracker_data = []
        
        # Add SIPP bonds
        if not sipp_ladder.empty:
            for _, bond in sipp_ladder.iterrows():
                tracker_data.append({
                    'ISIN': bond.get('isin', ''),
                    'Bond_Name': bond.get('bond_name', ''),
                    'Account_Type': 'SIPP',
                    'Maturity_Date': bond.get('maturity_date', ''),
                    'Coupon_Rate': bond.get('coupon', 0),
                    'Target_Allocation': bond.get('allocation', 0),
                    'Current_Price': '',  # To be filled manually
                    'YTM_Formula': f'=IF(G{len(tracker_data)+4}>0,YIELD(TODAY(),DATE({bond.get("maturity_date", "2027-01-01")[:4]},{bond.get("maturity_date", "2027-01-01")[5:7]},{bond.get("maturity_date", "2027-01-01")[8:10]}),{bond.get("coupon", 0)}%,G{len(tracker_data)+4},100,2,1)*100,"")',
                    'Price_Change': f'=IF(ROW()>4,IF(AND(G{len(tracker_data)+4}>0,G{len(tracker_data)+3}>0),(G{len(tracker_data)+4}-G{len(tracker_data)+3})/G{len(tracker_data)+3},""),"")',
                    'Status': f'=IF(G{len(tracker_data)+4}=0,"NOT PRICED",IF(H{len(tracker_data)+4}>4.5,"ATTRACTIVE",IF(H{len(tracker_data)+4}>4,"FAIR","EXPENSIVE")))',
                    'Action': f'=IF(G{len(tracker_data)+4}=0,"RESEARCH PRICE",IF(H{len(tracker_data)+4}>4.5,"CONSIDER BUYING",IF(H{len(tracker_data)+4}<3.5,"REVIEW","MONITOR")))'
                })
        
        # Add ISA bonds
        if not isa_ladder.empty:
            for _, bond in isa_ladder.iterrows():
                tracker_data.append({
                    'ISIN': bond.get('isin', ''),
                    'Bond_Name': bond.get('bond_name', ''),
                    'Account_Type': 'ISA',
                    'Maturity_Date': bond.get('maturity_date', ''),
                    'Coupon_Rate': bond.get('coupon', 0),
                    'Target_Allocation': bond.get('allocation', 0),
                    'Current_Price': '',  # To be filled manually
                    'YTM_Formula': f'=IF(G{len(tracker_data)+4}>0,YIELD(TODAY(),DATE({bond.get("maturity_date", "2027-01-01")[:4]},{bond.get("maturity_date", "2027-01-01")[5:7]},{bond.get("maturity_date", "2027-01-01")[8:10]}),{bond.get("coupon", 0)}%,G{len(tracker_data)+4},100,2,1)*100,"")',
                    'Price_Change': f'=IF(ROW()>4,IF(AND(G{len(tracker_data)+4}>0,G{len(tracker_data)+3}>0),(G{len(tracker_data)+4}-G{len(tracker_data)+3})/G{len(tracker_data)+3},""),"")',
                    'Status': f'=IF(G{len(tracker_data)+4}=0,"NOT PRICED",IF(H{len(tracker_data)+4}>4.5,"ATTRACTIVE",IF(H{len(tracker_data)+4}>4,"FAIR","EXPENSIVE")))',
                    'Action': f'=IF(G{len(tracker_data)+4}=0,"RESEARCH PRICE",IF(H{len(tracker_data)+4}>4.5,"CONSIDER BUYING",IF(H{len(tracker_data)+4}<3.5,"REVIEW","MONITOR")))'
                })
        
        # Create DataFrame
        df = pd.DataFrame(tracker_data)
        
        # Write to Excel with header
        worksheet = writer.book.create_sheet('Bond Price Tracker')
        
        # Add title and instructions
        worksheet['A1'] = 'Daily Bond Price Tracker'
        worksheet['A2'] = f'Last Updated: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A3'] = 'Instructions: Enter current prices in Column G, formulas will calculate YTM and status automatically'
        
        # Write headers
        headers = ['ISIN', 'Bond Name', 'Account', 'Maturity', 'Coupon %', 'Target £', 'Current Price', 'YTM %', 'Price Change %', 'Status', 'Action']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=4, column=col, value=header)
        
        # Write data (excluding formula columns initially)
        for row_idx, row_data in enumerate(tracker_data, 5):
            worksheet.cell(row=row_idx, column=1, value=row_data['ISIN'])
            worksheet.cell(row=row_idx, column=2, value=row_data['Bond_Name'])
            worksheet.cell(row=row_idx, column=3, value=row_data['Account_Type'])
            worksheet.cell(row=row_idx, column=4, value=row_data['Maturity_Date'])
            worksheet.cell(row=row_idx, column=5, value=row_data['Coupon_Rate'])
            worksheet.cell(row=row_idx, column=6, value=row_data['Target_Allocation'])
            # Column G (Current Price) left empty for manual entry
            worksheet.cell(row=row_idx, column=8, value=row_data['YTM_Formula'])
            worksheet.cell(row=row_idx, column=9, value=row_data['Price_Change'])
            worksheet.cell(row=row_idx, column=10, value=row_data['Status'])
            worksheet.cell(row=row_idx, column=11, value=row_data['Action'])
    
    def _create_purchase_log_sheet(self, writer: pd.ExcelWriter):
        """Create purchase logging sheet"""
        
        worksheet = writer.book.create_sheet('Purchase Log')
        
        # Title and instructions
        worksheet['A1'] = 'Bond Purchase Log'
        worksheet['A2'] = 'Record all bond purchases here for tracking and analysis'
        
        # Headers
        headers = ['Date', 'ISIN', 'Bond Name', 'Quantity £', 'Price Paid', 'Trading Fee', 'Total Cost', 'Account', 'Actual YTM %', 'Target YTM %', 'Variance', 'Status', 'Notes']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=4, column=col, value=header)
        
        # Sample formulas in first data row
        sample_row = 5
        worksheet.cell(row=sample_row, column=6, value=7.99)  # Default trading fee
        worksheet.cell(row=sample_row, column=7, value=f'=D{sample_row}+F{sample_row}')  # Total cost
        worksheet.cell(row=sample_row, column=9, value=f'=IF(E{sample_row}>0,YIELD(A{sample_row},DATE(2027,12,7),4.25%,E{sample_row},100,2,1)*100,"")')  # Actual YTM
        worksheet.cell(row=sample_row, column=11, value=f'=I{sample_row}-J{sample_row}')  # Variance
        worksheet.cell(row=sample_row, column=12, value=f'=IF(K{sample_row}>0.5,"BETTER THAN EXPECTED",IF(K{sample_row}<-0.5,"WORSE THAN EXPECTED","ON TARGET"))')  # Status
        
        # Add data validation for Account column
        from openpyxl.worksheet.datavalidation import DataValidation
        account_validation = DataValidation(type="list", formula1='"SIPP,ISA"')
        worksheet.add_data_validation(account_validation)
        account_validation.add(f'H{sample_row}:H100')
    
    def _create_quick_calculator_sheet(self, writer: pd.ExcelWriter):
        """Create quick bond calculation tools"""
        
        worksheet = writer.book.create_sheet('Quick Calculator')
        
        # Title
        worksheet['A1'] = 'Quick Bond Yield Calculator'
        
        # Simple YTM Calculator
        worksheet['A3'] = 'Single Bond Calculator'
        worksheet['A4'] = 'Current Price:'
        worksheet['B4'] = 98.5
        worksheet['A5'] = 'Face Value:'
        worksheet['B5'] = 100
        worksheet['A6'] = 'Coupon Rate (%):'
        worksheet['B6'] = 4.25
        worksheet['A7'] = 'Years to Maturity:'
        worksheet['B7'] = 2.5
        
        worksheet['A9'] = 'Results:'
        worksheet['A10'] = 'Yield to Maturity:'
        worksheet['B10'] = '=((B6*B5/100+(B5-B4)/B7)/((B5+B4)/2))*100'
        worksheet['A11'] = 'Current Yield:'
        worksheet['B11'] = '=(B6*B5/100)/B4*100'
        
        # Bond Comparison Tool
        worksheet['D1'] = 'Bond Comparison Tool'
        worksheet['D3'] = 'Bond A Price:'
        worksheet['E3'] = 102.0
        worksheet['F3'] = 'Bond B Price:'
        worksheet['G3'] = 96.5
        worksheet['D4'] = 'Bond A Coupon:'
        worksheet['E4'] = 4.25
        worksheet['F4'] = 'Bond B Coupon:'
        worksheet['G4'] = 3.5
        worksheet['D5'] = 'Bond A Maturity:'
        worksheet['E5'] = 3
        worksheet['F5'] = 'Bond B Maturity:'
        worksheet['G5'] = 4
        
        worksheet['D7'] = 'Bond A YTM:'
        worksheet['E7'] = '=((E4*100/100+(100-E3)/E5)/((100+E3)/2))*100'
        worksheet['F7'] = 'Bond B YTM:'
        worksheet['G7'] = '=((G4*100/100+(100-G3)/G5)/((100+G3)/2))*100'
        worksheet['D8'] = 'Better Choice:'
        worksheet['E8'] = '=IF(E7>G7,"Bond A","Bond B")'
        worksheet['F8'] = 'Yield Difference:'
        worksheet['G8'] = '=ABS(E7-G7)'
        
        # Interactive Investor Cost Calculator
        worksheet['A13'] = 'Interactive Investor Cost Calculator'
        worksheet['A14'] = 'Number of Bond Purchases:'
        worksheet['B14'] = 10
        worksheet['A15'] = 'Trading Fee per Purchase:'
        worksheet['B15'] = 7.99
        worksheet['A16'] = 'Monthly Platform Fee:'
        worksheet['B16'] = 12.99
        worksheet['A17'] = 'Annual Drawdown Fee:'
        worksheet['B17'] = 125
        
        worksheet['A19'] = 'Total Annual Cost:'
        worksheet['B19'] = '=B14*B15+B16*12+B17'
        worksheet['A20'] = 'Cost per £1000 invested:'
        worksheet['B20'] = '=B19/500'  # Assuming £500k portfolio
    
    def _create_portfolio_summary_sheet(self, writer: pd.ExcelWriter, 
                                      annual_data: List[Dict],
                                      sipp_ladder: pd.DataFrame, 
                                      isa_ladder: pd.DataFrame):
        """Create portfolio summary dashboard"""
        
        worksheet = writer.book.create_sheet('Portfolio Summary')
        
        # Title
        worksheet['A1'] = 'Portfolio Summary Dashboard'
        worksheet['A2'] = f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # SIPP Summary
        worksheet['A4'] = 'SIPP Portfolio Summary'
        worksheet['A5'] = 'Total Target Allocation:'
        worksheet['B5'] = f'=SUM(SIPP_Bonds[Target_Allocation])'
        worksheet['A6'] = 'Number of Bonds:'
        worksheet['B6'] = f'=COUNTA(SIPP_Bonds[ISIN])'
        worksheet['A7'] = 'Average Target Yield:'
        worksheet['B7'] = f'=AVERAGE(SIPP_Bonds[Estimated_YTM])'
        worksheet['A8'] = 'Projected Annual Income:'
        worksheet['B8'] = f'=SUM(SIPP_Bonds[Annual_Income])'
        
        # ISA Summary
        worksheet['D4'] = 'ISA Portfolio Summary'
        worksheet['D5'] = 'Total Target Allocation:'
        worksheet['E5'] = f'=SUM(ISA_Bonds[Target_Allocation])'
        worksheet['D6'] = 'Number of Bonds:'
        worksheet['E6'] = f'=COUNTA(ISA_Bonds[ISIN])'
        worksheet['D7'] = 'Average Target Yield:'
        worksheet['E7'] = f'=AVERAGE(ISA_Bonds[Estimated_YTM])'
        worksheet['D8'] = 'Projected Annual Income:'
        worksheet['E8'] = f'=SUM(ISA_Bonds[Annual_Income])'
        
        # Combined Portfolio
        worksheet['A10'] = 'Combined Portfolio'
        worksheet['A11'] = 'Total Portfolio Value:'
        worksheet['B11'] = f'=B5+E5'
        worksheet['A12'] = 'Total Annual Income:'
        worksheet['B12'] = f'=B8+E8'
        worksheet['A13'] = 'Portfolio Yield:'
        worksheet['B13'] = f'=B12/B11'
        
        # Risk Indicators
        worksheet['A15'] = 'Risk Indicators'
        worksheet['A16'] = 'Bonds > 5% YTM:'
        worksheet['B16'] = f'=COUNTIF(\'Bond Price Tracker\'!H:H,">5")'
        worksheet['A17'] = 'Bonds < 3% YTM:'
        worksheet['B17'] = f'=COUNTIF(\'Bond Price Tracker\'!H:H,"<3")'
        worksheet['A18'] = 'Bonds Needing Review:'
        worksheet['B18'] = f'=COUNTIF(\'Bond Price Tracker\'!J:J,"REVIEW")'
        worksheet['A19'] = 'Action Items:'
        worksheet['B19'] = f'=COUNTIF(\'Bond Price Tracker\'!K:K,"CONSIDER BUYING")'
        
        # Performance tracking (if annual data available)
        if annual_data:
            worksheet['D10'] = 'Strategy Performance'
            worksheet['D11'] = 'Year 1 Target Income:'
            worksheet['E11'] = annual_data[0].get('target_income', 0)
            worksheet['D12'] = 'Year 1 Projected Income:'
            worksheet['E12'] = annual_data[0].get('total_net_income', 0)
            worksheet['D13'] = 'Income Achievement:'
            worksheet['E13'] = f'=E12/E11'
            worksheet['D14'] = 'Estimated Tax Rate:'
            worksheet['E14'] = annual_data[0].get('effective_tax_rate', 0) / 100
    
    def _create_risk_dashboard_sheet(self, writer: pd.ExcelWriter, 
                                   sipp_ladder: pd.DataFrame, 
                                   isa_ladder: pd.DataFrame):
        """Create risk monitoring dashboard"""
        
        worksheet = writer.book.create_sheet('Risk Dashboard')
        
        # Title
        worksheet['A1'] = 'Risk Monitoring Dashboard'
        
        # Credit Risk Analysis
        worksheet['A3'] = 'Credit Risk Analysis'
        worksheet['A4'] = 'Government Bonds (SIPP):'
        worksheet['B4'] = f'=COUNTIF(\'Bond Price Tracker\'!C:C,"SIPP")'
        worksheet['A5'] = 'Corporate Bonds (ISA):'
        worksheet['B5'] = f'=COUNTIF(\'Bond Price Tracker\'!C:C,"ISA")'
        worksheet['A6'] = 'Investment Grade %:'
        worksheet['B6'] = '=100%'  # Assuming all bonds are investment grade
        
        # Maturity Risk Analysis
        worksheet['A8'] = 'Maturity Risk Analysis'
        worksheet['A9'] = 'Bonds Maturing 2025:'
        worksheet['B9'] = f'=COUNTIF(\'Bond Price Tracker\'!D:D,"*2025*")'
        worksheet['A10'] = 'Bonds Maturing 2026:'
        worksheet['B10'] = f'=COUNTIF(\'Bond Price Tracker\'!D:D,"*2026*")'
        worksheet['A11'] = 'Bonds Maturing 2027:'
        worksheet['B11'] = f'=COUNTIF(\'Bond Price Tracker\'!D:D,"*2027*")'
        worksheet['A12'] = 'Bonds Maturing 2028+:'
        worksheet['B12'] = f'=COUNTIF(\'Bond Price Tracker\'!D:D,"*2028*")+COUNTIF(\'Bond Price Tracker\'!D:D,"*2029*")+COUNTIF(\'Bond Price Tracker\'!D:D,"*2030*")'
        
        # Yield Risk Analysis
        worksheet['A14'] = 'Yield Risk Analysis'
        worksheet['A15'] = 'High Yield Bonds (>5%):'
        worksheet['B15'] = f'=COUNTIF(\'Bond Price Tracker\'!H:H,">5")'
        worksheet['A16'] = 'Medium Yield Bonds (3-5%):'
        worksheet['B16'] = f'=COUNTIFS(\'Bond Price Tracker\'!H:H,">=3",\'Bond Price Tracker\'!H:H,"<=5")'
        worksheet['A17'] = 'Low Yield Bonds (<3%):'
        worksheet['B17'] = f'=COUNTIF(\'Bond Price Tracker\'!H:H,"<3")'
        
        # Sector Concentration (for ISA)
        worksheet['D3'] = 'Sector Concentration (ISA)'
        worksheet['D4'] = 'Utilities:'
        worksheet['E4'] = '=COUNTIF(Bond_Sectors,"Utilities")/COUNTA(Bond_Sectors)'
        worksheet['D5'] = 'Telecommunications:'
        worksheet['E5'] = '=COUNTIF(Bond_Sectors,"Telecommunications")/COUNTA(Bond_Sectors)'
        worksheet['D6'] = 'Banking:'
        worksheet['E6'] = '=COUNTIF(Bond_Sectors,"Banking")/COUNTA(Bond_Sectors)'
        worksheet['D7'] = 'Insurance:'
        worksheet['E7'] = '=COUNTIF(Bond_Sectors,"Insurance")/COUNTA(Bond_Sectors)'
        worksheet['D8'] = 'Other:'
        worksheet['E8'] = '=1-(E4+E5+E6+E7)'
        
        # Alert System
        worksheet['A19'] = 'Alert System'
        worksheet['A20'] = 'Concentration Risk:'
        worksheet['B20'] = '=IF(MAX(E4:E8)>0.4,"HIGH","OK")'
        worksheet['A21'] = 'Yield Risk:'
        worksheet['B21'] = '=IF(B17>B15,"HIGH","OK")'
        worksheet['A22'] = 'Maturity Risk:'
        worksheet['B22'] = '=IF(B9>B12,"HIGH","OK")'
    
    def _create_yield_analysis_sheet(self, writer: pd.ExcelWriter, 
                                   sipp_ladder: pd.DataFrame, 
                                   isa_ladder: pd.DataFrame):
        """Create yield analysis and comparison sheet"""
        
        worksheet = writer.book.create_sheet('Yield Analysis')
        
        # Title
        worksheet['A1'] = 'Yield Analysis & Market Comparison'
        
        # Current Market Benchmarks
        worksheet['A3'] = 'Market Benchmarks'
        worksheet['A4'] = 'UK 10-Year Gilt:'
        worksheet['B4'] = '4.56%'  # Current as of search results
        worksheet['A5'] = 'UK 5-Year Gilt:'
        worksheet['B5'] = '4.20%'
        worksheet['A6'] = 'UK 2-Year Gilt:'
        worksheet['B6'] = '3.80%'
        worksheet['A7'] = 'Investment Grade Corporate:'
        worksheet['B7'] = '4.80%'
        
        # Portfolio Yield Analysis
        worksheet['A9'] = 'Portfolio Yield Analysis'
        worksheet['A10'] = 'SIPP Average Yield:'
        worksheet['B10'] = f'=AVERAGEIF(\'Bond Price Tracker\'!C:C,"SIPP",\'Bond Price Tracker\'!H:H)'
        worksheet['A11'] = 'ISA Average Yield:'
        worksheet['B11'] = f'=AVERAGEIF(\'Bond Price Tracker\'!C:C,"ISA",\'Bond Price Tracker\'!H:H)'
        worksheet['A12'] = 'Portfolio Weighted Yield:'
        worksheet['B12'] = f'=SUMPRODUCT(\'Bond Price Tracker\'!F:F,\'Bond Price Tracker\'!H:H)/SUM(\'Bond Price Tracker\'!F:F)'
        
        # Yield vs Benchmark Analysis
        worksheet['A14'] = 'Yield vs Benchmark'
        worksheet['A15'] = 'SIPP vs 10Y Gilt:'
        worksheet['B15'] = f'=B10-B4'
        worksheet['A16'] = 'ISA vs Corp Avg:'
        worksheet['B16'] = f'=B11-B7'
        worksheet['A17'] = 'Portfolio Premium:'
        worksheet['B17'] = f'=B12-B4'
        
        # Yield Sensitivity Analysis
        worksheet['D3'] = 'Yield Sensitivity Analysis'
        worksheet['D4'] = 'If yields rise by 1%:'
        worksheet['E4'] = 'Price Impact'
        worksheet['D5'] = 'Short-term bonds (1-3y):'
        worksheet['E5'] = '-2.5%'
        worksheet['D6'] = 'Medium-term bonds (3-7y):'
        worksheet['E6'] = '-5.0%'
        worksheet['D7'] = 'Long-term bonds (7y+):'
        worksheet['E7'] = '-7.5%'
        
        # Income Projections
        worksheet['D9'] = 'Income Projections'
        worksheet['D10'] = 'Monthly Income (Current):'
        worksheet['E10'] = f'=\'Portfolio Summary\'!B12/12'
        worksheet['D11'] = 'Monthly Income (If yields +1%):'
        worksheet['E11'] = f'=E10*1.01'
        worksheet['D12'] = 'Monthly Income (If yields -1%):'
        worksheet['E12'] = f'=E10*0.99'
    
    def _create_implementation_checklist_sheet(self, writer: pd.ExcelWriter):
        """Create implementation checklist and timeline"""
        
        worksheet = writer.book.create_sheet('Implementation Checklist')
        
        # Title
        worksheet['A1'] = 'Implementation Checklist & Timeline'
        
        # Pre-Implementation
        worksheet['A3'] = 'Pre-Implementation (Before Account Opening)'
        checklist_items = [
            'Research Interactive Investor platform',
            'Gather required documents (ID, bank statements)',
            'Calculate target allocations using Streamlit app',
            'Download and review this Excel workbook',
            'Set up spreadsheet for daily use'
        ]
        
        for i, item in enumerate(checklist_items, 4):
            worksheet[f'A{i}'] = item
            worksheet[f'B{i}'] = '☐'  # Checkbox
            worksheet[f'C{i}'] = ''   # Completion date
            worksheet[f'D{i}'] = ''   # Notes
        
        # Week 1
        worksheet['A10'] = 'Week 1: Account Setup'
        week1_items = [
            'Open Interactive Investor SIPP account',
            'Open Interactive Investor ISA account',
            'Set up Direct Debit for monthly fees',
            'Initiate fund transfers from existing accounts',
            'Research current bond prices using target ISINs'
        ]
        
        for i, item in enumerate(week1_items, 11):
            worksheet[f'A{i}'] = item
            worksheet[f'B{i}'] = '☐'
            worksheet[f'C{i}'] = ''
            worksheet[f'D{i}'] = ''
        
        # Week 2-3
        worksheet['A17'] = 'Week 2-3: First Purchases'
        week2_items = [
            'Purchase first UK Gilt (highest priority)',
            'Purchase second UK Gilt',
            'Start ISA corporate bond purchases',
            'Update purchase log after each transaction',
            'Monitor yield changes daily'
        ]
        
        for i, item in enumerate(week2_items, 18):
            worksheet[f'A{i}'] = item
            worksheet[f'B{i}'] = '☐'
            worksheet[f'C{i}'] = ''
            worksheet[f'D{i}'] = ''
        
        # Week 4
        worksheet['A24'] = 'Week 4: Complete Initial Allocation'
        week4_items = [
            'Complete remaining bond purchases',
            'Verify all allocations match targets',
            'Set up monitoring routine',
            'Review and adjust strategy if needed',
            'Plan first reinvestment (if applicable)'
        ]
        
        for i, item in enumerate(week4_items, 25):
            worksheet[f'A{i}'] = item
            worksheet[f'B{i}'] = '☐'
            worksheet[f'C{i}'] = ''
            worksheet[f'D{i}'] = ''
        
        # Ongoing Tasks
        worksheet['A31'] = 'Ongoing Tasks'
        ongoing_items = [
            'Daily: Update bond prices in tracker',
            'Weekly: Review portfolio performance',
            'Monthly: Analyze yield trends',
            'Quarterly: Review strategy and rebalance',
            'As needed: Reinvest maturing bonds'
        ]
        
        for i, item in enumerate(ongoing_items, 32):
            worksheet[f'A{i}'] = item
            worksheet[f'B{i}'] = '☐'
            worksheet[f'C{i}'] = ''
            worksheet[f'D{i}'] = ''
        
        # Headers
        worksheet['A2'] = 'Task'
        worksheet['B2'] = 'Done'
        worksheet['C2'] = 'Date'
        worksheet['D2'] = 'Notes'
    
    def _create_sipp_ladder_sheet(self, writer: pd.ExcelWriter, sipp_ladder: pd.DataFrame):
        """Create enhanced SIPP ladder sheet"""
        sipp_enhanced = sipp_ladder.copy()
        
        # Add practical columns
        sipp_enhanced['Bonds_Needed'] = sipp_enhanced['allocation'] / 100  # Assuming £100 bonds
        sipp_enhanced['Trading_Cost'] = 7.99
        sipp_enhanced['Total_Cost'] = sipp_enhanced['allocation'] + sipp_enhanced['Trading_Cost']
        sipp_enhanced['Purchase_Priority'] = range(1, len(sipp_enhanced) + 1)
        sipp_enhanced['Purchase_Status'] = 'PENDING'
        sipp_enhanced['Actual_Price'] = ''
        sipp_enhanced['Actual_YTM'] = ''
        sipp_enhanced['Purchase_Date'] = ''
        
        sipp_enhanced.to_excel(writer, sheet_name='SIPP Bond Ladder Enhanced', index=False)
    
    def _create_isa_ladder_sheet(self, writer: pd.ExcelWriter, isa_ladder: pd.DataFrame):
        """Create enhanced ISA ladder sheet"""
        isa_enhanced = isa_ladder.copy()
        
        # Add practical columns
        isa_enhanced['Bonds_Needed'] = isa_enhanced['allocation'] / 100  # Assuming £100 bonds
        isa_enhanced['Trading_Cost'] = 7.99
        isa_enhanced['Total_Cost'] = isa_enhanced['allocation'] + isa_enhanced['Trading_Cost']
        isa_enhanced['Purchase_Priority'] = range(1, len(isa_enhanced) + 1)
        isa_enhanced['Purchase_Status'] = 'PENDING'
        isa_enhanced['Actual_Price'] = ''
        isa_enhanced['Actual_YTM'] = ''
        isa_enhanced['Purchase_Date'] = ''
        
        isa_enhanced.to_excel(writer, sheet_name='ISA Bond Ladder Enhanced', index=False)
    
    def _create_annual_analysis_sheet(self, writer: pd.ExcelWriter, annual_data: List[Dict]):
        """Create enhanced annual analysis sheet"""
        if not annual_data:
            return
            
        df = pd.DataFrame(annual_data)
        
        # Add additional calculated columns
        df['Bond_Income_Total'] = df.get('sipp_bond_income', 0) + df.get('isa_bond_income', 0)
        df['Bond_Income_Percent'] = df['Bond_Income_Total'] / df['total_net_income'] * 100
        df['Tax_Efficiency'] = 100 - df['effective_tax_rate']
        df['Portfolio_Drawdown_Rate'] = (df.get('drawdown_tax_free', 0) + 
                                       df.get('drawdown_isa', 0) + 
                                       df.get('drawdown_taxable', 0)) / df['total_remaining_pots'] * 100
        
        df.to_excel(writer, sheet_name='Annual Analysis Enhanced', index=False)
    
    def _create_scenario_analysis_sheet(self, writer: pd.ExcelWriter, scenario_results: Dict):
        """Create scenario analysis sheet"""
        scenario_data = []
        for scenario_name, results in scenario_results.items():
            scenario_data.append({
                'Scenario': scenario_name,
                'Description': results.get('description', ''),
                'Final_Portfolio_Value': results.get('final_pot', 0),
                'Average_Net_Income': results.get('avg_net_income', 0),
                'Average_Tax_Rate': results.get('avg_tax_rate', 0),
                'Sustainability_Score': 'High' if results.get('final_pot', 0) > 100000 else 'Medium' if results.get('final_pot', 0) > 50000 else 'Low'
            })
        
        df = pd.DataFrame(scenario_data)
        df.to_excel(writer, sheet_name='Scenario Analysis', index=False)
    
    def _create_monte_carlo_sheet(self, writer: pd.ExcelWriter, monte_carlo_results: List):
        """Create Monte Carlo results sheet"""
        df = pd.DataFrame(monte_carlo_results)
        
        # Add percentile analysis
        percentiles = [10, 25, 50, 75, 90, 95, 99]
        percentile_data = []
        
        for p in percentiles:
            percentile_data.append({
                'Percentile': f'{p}th',
                'Final_Portfolio_Value': np.percentile(df['final_pot'], p),
                'Average_Income': np.percentile(df['avg_income'], p),
                'Average_Tax_Rate': np.percentile(df['avg_tax_rate'], p)
            })
        
        df_percentiles = pd.DataFrame(percentile_data)
        
        # Write both datasets
        df.to_excel(writer, sheet_name='Monte Carlo Results', index=False)
        df_percentiles.to_excel(writer, sheet_name='Monte Carlo Percentiles', index=False)
    
    def _apply_formatting(self, writer: pd.ExcelWriter):
        """Apply formatting to all sheets"""
        workbook = writer.book
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                if cell.value:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color=self.colors['header'], 
                                          end_color=self.colors['header'], 
                                          fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add conditional formatting for specific sheets
            if sheet_name == 'Bond Price Tracker':
                # Color scale for YTM column (assuming column H)
                ytm_range = f'H5:H{worksheet.max_row}'
                rule = ColorScaleRule(start_type='min', start_color='FF0000',  # Red
                                    mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow
                                    end_type='max', end_color='00FF00')  # Green
                worksheet.conditional_formatting.add(ytm_range, rule)
                
                # Status column formatting
                status_range = f'J5:J{worksheet.max_row}'
                green_rule = CellIsRule(operator='equal', formula=['"ATTRACTIVE"'], 
                                      fill=PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'))
                red_rule = CellIsRule(operator='equal', formula=['"EXPENSIVE"'], 
                                    fill=PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'))
                worksheet.conditional_formatting.add(status_range, green_rule)
                worksheet.conditional_formatting.add(status_range, red_rule)


# Usage example for integration with main app
def create_enhanced_excel_export_wrapper(annual_data, sipp_ladder, isa_ladder, 
                                        scenario_results=None, monte_carlo_results=None):
    """
    Wrapper function to integrate with existing main app
    """
    try:
        exporter = EnhancedExcelExporter()
        return exporter.create_complete_excel_export(
            annual_data, sipp_ladder, isa_ladder, 
            scenario_results, monte_carlo_results
        )
    except Exception as e:
        logging.error(f"Enhanced Excel export failed: {str(e)}")
        return None
